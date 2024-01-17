from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from mrr import calculate_simple_mrr
from churn import calculate_churn_rate
import pandas as pd
import io
import csv
from openpyxl import load_workbook
   
app = FastAPI()

origins = [
   'http://localhost:3000'
]

app.add_middleware(
  CORSMiddleware,
  allow_origins=origins,
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"Hello": "World"}

@app.get("/items/{item_id}")
def read_item(item_id: int, query_param: str = None):
    return {"item_id": item_id, "query_param": query_param}


## TODO: protect API if a KEY
## TODO: unit test for this route

@app.post("/upload-file/")
async def upload_file(file: UploadFile = File(...)):
  try:
    extension = file.filename.split('.')[-1]
    data = await file.read();
  
    if extension == 'csv':
      mrr_result = calculate_simple_mrr(io.StringIO(data.decode('utf-8')))
      churn_rate_result = calculate_churn_rate(io.StringIO(data.decode('utf-8')))

      result = {"mrr": mrr_result, "churn_rate": churn_rate_result}
      return JSONResponse(content=result)
    
    elif extension == 'xlsx':
      workbook = load_workbook(io.BytesIO(data))
      sheet = workbook.active
      csv_data = io.StringIO()
      csv_writer = csv.writer(csv_data)

      for row in sheet.iter_rows(min_row=1, values_only=True):
        csv_writer.writerow(row)

      mrr_result = calculate_simple_mrr(io.StringIO(csv_data.getvalue()))
      churn_rate_result = calculate_churn_rate(io.StringIO(csv_data.getvalue()))
      result = {"mrr": mrr_result, "churn_rate": churn_rate_result}
      return JSONResponse(content=result)


  except Exception as e:
      raise HTTPException(status_code=500, detail=f"Falha ao processar o arquivo: {str(e)}")